import openpyxl
from openpyxl.styles import PatternFill

# Charger le fichier Excel
workbook = openpyxl.load_workbook('Mask_Ref.xlsx')
sheet = workbook.active

# Définir le style de remplissage pour les cellules à colorer
fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Parcourir les cellules et appliquer le style si la valeur est inférieure à 6
for row in sheet.iter_rows():
    for cell in row:
        if isinstance(cell.value, (int, float)) and cell.value < 6:
            cell.fill = fill

# Sauvegarder le fichier modifié
workbook.save('votre_fichier_colore.xlsx')
