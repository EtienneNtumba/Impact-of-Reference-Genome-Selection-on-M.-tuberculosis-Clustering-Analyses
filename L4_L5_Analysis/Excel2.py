import openpyxl
from openpyxl.styles import PatternFill

# Charger le fichier Excel
workbook = openpyxl.load_workbook('votre_fichier_colore.xlsx')
sheet = workbook.active

# Définir les styles de remplissage pour les cellules
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

# Parcourir les cellules et appliquer les styles selon les conditions
for row in sheet.iter_rows():
    for cell in row:
        if isinstance(cell.value, (int, float)):
            if 0 < cell.value < 6:
                cell.fill = yellow_fill
            elif 7 <= cell.value <= 12:
                cell.fill = orange_fill

# Sauvegarder le fichier modifié
workbook.save('votre_fichier_colore1.xlsx')
